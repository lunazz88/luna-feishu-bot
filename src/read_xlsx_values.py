import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def cell_value(value):
    if value is None:
        return ""
    return value


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    if len(sys.argv) < 2:
        raise SystemExit("usage: read_xlsx_values.py <xlsx_path>")

    workbook_path = Path(sys.argv[1])
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    # Some crawler files incorrectly declare the worksheet dimension as A1.
    # Force openpyxl to scan rows instead of trusting that stale metadata.
    if hasattr(sheet, "reset_dimensions"):
        sheet.reset_dimensions()
    rows = []
    for row in sheet.iter_rows(values_only=True):
        values = [cell_value(value) for value in row]
        while values and values[-1] == "":
            values.pop()
        rows.append(values)
    print(json.dumps(rows, ensure_ascii=False))


if __name__ == "__main__":
    main()
